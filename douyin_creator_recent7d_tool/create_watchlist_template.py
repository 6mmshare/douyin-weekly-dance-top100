#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).resolve().parent / "creators_watchlist.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "博主名单"
headers = [
    "启用", "备注", "抖音名称", "抖音账号（精确）",
    "粉丝数（原文）", "粉丝数（数值）",
    "获赞数（原文）", "获赞数（数值）", "原始博主信息", "来源行",
]
ws.append(headers)
ws.append(["是", "示例，使用前请删除", "示例账号", "example_id", "10万", 100000, "100万", 1000000, "", 2])
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "A2"
widths = [9, 22, 22, 25, 15, 16, 15, 16, 42, 10]
for index, width in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + index)].width = width
wb.save(OUT)
print(f"已生成：{OUT}")
