#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import dataclasses
import hashlib
import html
import json
import math
import os
import re
import subprocess
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "creator_recent7d_config.yaml"
CREATORS_PATH = APP_DIR / "creators_watchlist.xlsx"
DATA_DIR = APP_DIR / "data"
PROFILE_DIR = APP_DIR.parent / "data" / "profiles" / "douyin"
CHECKPOINT_PATH = DATA_DIR / "creator_recent7d_checkpoint.json"
LATEST_PATH = DATA_DIR / "latest_collection.txt"
LOCK_PATH = DATA_DIR / "creator_recent7d.lock"
LOG_DIR = APP_DIR / "logs"
COUNT_RE = re.compile(r"(-?\d+(?:\.\d+)?)\s*([万亿千kKmMbB]?)")
ACCOUNT_RE = re.compile(r"抖音号\s*[:：]\s*([^\s\r\n]+)", re.I)


def ensure_dirs():
    for path in (DATA_DIR, PROFILE_DIR, LOG_DIR):
        path.mkdir(parents=True, exist_ok=True)


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm_account(value: Any) -> str:
    return norm(value).lstrip("@").rstrip("，。；;、")


def account_key(value: Any) -> str:
    return norm_account(value).casefold()


def parse_count(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return max(0, int(value))
        except Exception:
            return 0
    text = str(value).strip().replace(",", "").replace("+", "")
    try:
        return max(0, int(float(text)))
    except Exception:
        pass
    match = COUNT_RE.search(text)
    if not match:
        return 0
    number = float(match.group(1))
    unit = match.group(2).lower()
    return max(0, int(number * {"": 1, "千": 1e3, "k": 1e3, "万": 1e4, "m": 1e6, "亿": 1e8, "b": 1e9}.get(unit, 1)))


def parse_dt(value: Any) -> Optional[datetime]:
    if value in (None, "", 0, "0"):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, (int, float)) or (isinstance(value, str) and value.strip().isdigit()):
        try:
            number = float(value)
            if number > 10_000_000_000:
                number /= 1000
            if number < 946684800:
                return None
            return datetime.fromtimestamp(number, tz=timezone.utc)
        except Exception:
            return None
    text = str(value).strip().replace("Z", "+00:00")
    for parser in (datetime.fromisoformat, lambda x: datetime.strptime(x, "%Y-%m-%d %H:%M:%S"), lambda x: datetime.strptime(x, "%Y-%m-%d")):
        try:
            result = parser(text)
            return result if result.tzinfo else result.replace(tzinfo=timezone.utc)
        except Exception:
            pass
    return None


def deep_get(obj: Mapping[str, Any], *paths: str, default=None):
    for path in paths:
        current: Any = obj
        for part in path.split("."):
            if not isinstance(current, Mapping) or part not in current:
                current = None
                break
            current = current[part]
        if current not in (None, ""):
            return current
    return default


def walk_dicts(obj: Any) -> Iterator[dict[str, Any]]:
    if isinstance(obj, dict):
        yield obj
        for value in obj.values():
            yield from walk_dicts(value)
    elif isinstance(obj, list):
        for value in obj:
            yield from walk_dicts(value)


def first_url(value: Any) -> str:
    if isinstance(value, str):
        return value if value.startswith("http") else ""
    if isinstance(value, list):
        for item in value:
            url = first_url(item)
            if url:
                return url
    if isinstance(value, Mapping):
        for key in ("url_list", "urlList", "urls", "url", "src"):
            if key in value:
                url = first_url(value[key])
                if url:
                    return url
    return ""


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8-sig")) or {}


def resolve_tz(name: str):
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        if name.startswith("Asia/"):
            return timezone(timedelta(hours=8), name=name)
        raise


@dataclass
class Window:
    start: datetime
    end: datetime
    mode: str

    @property
    def label(self):
        return f"{self.start:%Y-%m-%d %H:%M:%S} 至 {self.end:%Y-%m-%d %H:%M:%S}（左闭右开）"

    @property
    def slug(self):
        return f"{self.start:%Y-%m-%d_%H%M}__{self.end:%Y-%m-%d_%H%M}"

    def to_dict(self):
        return {"start": self.start.isoformat(), "end": self.end.isoformat(), "mode": self.mode, "label": self.label, "slug": self.slug}

    @classmethod
    def from_dict(cls, raw):
        return cls(parse_dt(raw["start"]), parse_dt(raw["end"]), raw.get("mode", "rolling_last_7_days"))


def compute_window(config, now=None):
    tz = resolve_tz(str(config.get("timezone", "Asia/Shanghai")))
    current = (now or datetime.now(tz)).astimezone(tz)
    mode = str(config.get("window_mode", "rolling_last_7_days"))
    if mode == "rolling_last_7_days":
        end = current.replace(microsecond=0)
        start = end - timedelta(days=7)
    elif mode == "previous_7_complete_days":
        end = current.replace(hour=0, minute=0, second=0, microsecond=0)
        start = end - timedelta(days=7)
    elif mode == "previous_calendar_week":
        today = current.replace(hour=0, minute=0, second=0, microsecond=0)
        end = today - timedelta(days=today.weekday())
        start = end - timedelta(days=7)
    else:
        raise ValueError(f"不支持 window_mode：{mode}")
    return Window(start, end, mode), tz


def in_window(dt, window, tz):
    if not dt:
        return False
    value = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    return window.start <= value.astimezone(tz) < window.end


@dataclass
class Creator:
    row_number: int
    enabled: bool
    remark: str
    name: str
    account: str
    expected_followers: int = 0
    expected_likes: int = 0
    raw_info: str = ""
    homepage: str = ""
    sec_uid: str = ""
    uid: str = ""
    resolved_name: str = ""
    resolved_account: str = ""
    current_followers: int = 0
    status: str = "pending"
    note: str = ""

    @property
    def key(self):
        return account_key(self.account)

    def to_dict(self):
        return dataclasses.asdict(self)

    @classmethod
    def from_dict(cls, raw):
        fields = {item.name for item in dataclasses.fields(cls)}
        return cls(**{key: value for key, value in raw.items() if key in fields})


@dataclass
class Video:
    video_id: str
    url: str
    creator_name: str
    creator_account: str
    creator_homepage: str
    author_name: str = ""
    author_unique_id: str = ""
    author_short_id: str = ""
    author_uid: str = ""
    author_sec_uid: str = ""
    title: str = ""
    create_time: Optional[datetime] = None
    likes: int = 0
    comments: int = 0
    shares: int = 0
    favorites: int = 0
    views: int = 0
    music: str = ""
    thumbnail: str = ""
    source: str = "profile_api"
    captured_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    notes: list[str] = field(default_factory=list)
    likes_per_hour: float = 0.0
    like_percentile: float = 0.0
    comment_percentile: float = 0.0
    share_percentile: float = 0.0
    favorite_percentile: float = 0.0
    velocity_percentile: float = 0.0
    final_score: float = 0.0
    rank: int = 0

    def to_dict(self, tz=None):
        result = dataclasses.asdict(self)
        if self.create_time:
            result["create_time"] = self.create_time.astimezone(tz).isoformat() if tz else self.create_time.isoformat()
        result["captured_at"] = self.captured_at.isoformat()
        return result

    @classmethod
    def from_dict(cls, raw):
        fields = {item.name for item in dataclasses.fields(cls)}
        data = {key: value for key, value in raw.items() if key in fields}
        data["create_time"] = parse_dt(data.get("create_time"))
        data["captured_at"] = parse_dt(data.get("captured_at")) or datetime.now(timezone.utc)
        data["notes"] = list(data.get("notes") or [])
        return cls(**data)


def read_creators(path: Path = CREATORS_PATH):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    aliases = {"enabled": ("启用",), "remark": ("备注",), "raw": ("原始博主信息", "博主信息"), "name": ("抖音名称", "账号名称", "昵称"), "account": ("抖音账号", "抖音号", "账号"), "followers": ("粉丝数",), "likes": ("获赞数",)}
    header_row, cols = None, {}
    for row in range(1, min(ws.max_row, 12) + 1):
        found = {}
        for col in range(1, ws.max_column + 1):
            value = norm(ws.cell(row, col).value)
            for key, names in aliases.items():
                if any(value == name or value.startswith(name + "（") for name in names):
                    found.setdefault(key, col)
        if "name" in found and "account" in found:
            header_row, cols = row, found
            break
    if not header_row:
        raise ValueError("Excel中未找到抖音名称和抖音账号表头")
    creators, seen = [], set()
    for row in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row, cols.get("raw", 0)).value if cols.get("raw") else ""
        match = ACCOUNT_RE.search(str(raw or ""))
        account = norm_account(match.group(1) if match else ws.cell(row, cols["account"]).value)
        if not account or account_key(account) in seen:
            continue
        enabled = norm(ws.cell(row, cols.get("enabled", 0)).value if cols.get("enabled") else "是").casefold() not in {"否", "false", "0", "no", "n"}
        if not enabled:
            continue
        seen.add(account_key(account))
        creators.append(Creator(row, True, norm(ws.cell(row, cols.get("remark", 0)).value if cols.get("remark") else ""), norm(ws.cell(row, cols["name"]).value), account, parse_count(ws.cell(row, cols.get("followers", 0)).value if cols.get("followers") else 0), parse_count(ws.cell(row, cols.get("likes", 0)).value if cols.get("likes") else 0), norm(raw)))
    wb.close()
    return creators


def fingerprint(creators: Iterable[Creator], config):
    payload = {"creators": [(c.name, c.account) for c in creators], "window_mode": config.get("window_mode"), "collector": config.get("collector", {})}
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode()).hexdigest()[:24]


def atomic_json(path: Path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, path)


def load_json(path: Path, default=None):
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default


class Lock:
    def __init__(self, stale_hours=12):
        self.stale = stale_hours * 3600

    def __enter__(self):
        ensure_dirs()
        if LOCK_PATH.exists() and time.time() - LOCK_PATH.stat().st_mtime < self.stale:
            raise RuntimeError("已有指定博主采集任务运行")
        atomic_json(LOCK_PATH, {"pid": os.getpid(), "time": time.time()})
        return self

    def __exit__(self, *_):
        LOCK_PATH.unlink(missing_ok=True)


def percentile(values):
    if not values:
        return []
    if len(values) == 1:
        return [100.0]
    ordered = sorted(enumerate(values), key=lambda x: x[1])
    result = [0.0] * len(values)
    for rank, (index, _) in enumerate(ordered):
        result[index] = rank / (len(values) - 1) * 100
    return result


def rank_videos(videos, window, tz, config):
    eligible = [v for v in videos if in_window(v.create_time, window, tz)]
    for video in eligible:
        hours = max((window.end - video.create_time.astimezone(tz)).total_seconds() / 3600, 1)
        video.likes_per_hour = video.likes / hours
    fields = ("likes", "comments", "shares", "favorites", "likes_per_hour")
    sets = [percentile([math.log1p(getattr(v, field)) for v in eligible]) for field in fields]
    weights = [float(config.get("ranking", {}).get(name, default)) for name, default in (("likes", .35), ("comments", .10), ("shares", .20), ("favorites", .15), ("likes_velocity", .20))]
    total = sum(weights) or 1
    for i, video in enumerate(eligible):
        video.like_percentile, video.comment_percentile, video.share_percentile, video.favorite_percentile, video.velocity_percentile = [items[i] for items in sets]
        video.final_score = round(sum(items[i] * weight for items, weight in zip(sets, weights)) / total, 2)
    eligible.sort(key=lambda v: (v.final_score, v.likes, v.shares, v.favorites, v.comments), reverse=True)
    for rank, video in enumerate(eligible, 1):
        video.rank = rank
    return eligible


def _row(video, tz):
    return {"rank": video.rank, "score": video.final_score, "creator_name": video.creator_name, "creator_account": video.creator_account, "homepage": video.creator_homepage, "publish_time": video.create_time.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S") if video.create_time else "", "title": video.title, "url": video.url, "likes": video.likes, "comments": video.comments, "shares": video.shares, "favorites": video.favorites, "views": video.views, "music": video.music, "video_id": video.video_id, "author_account": video.author_unique_id or video.author_short_id, "source": video.source, "notes": "；".join(video.notes), "likes_per_hour": round(video.likes_per_hour, 2), "like_pct": round(video.like_percentile, 2), "comment_pct": round(video.comment_percentile, 2), "share_pct": round(video.share_percentile, 2), "favorite_pct": round(video.favorite_percentile, 2), "velocity_pct": round(video.velocity_percentile, 2)}


COLLECT_COLS = [("博主名称", "creator_name"), ("抖音账号", "creator_account"), ("博主主页", "homepage"), ("发布时间", "publish_time"), ("视频标题", "title"), ("视频链接", "url"), ("点赞", "likes"), ("评论", "comments"), ("分享", "shares"), ("收藏", "favorites"), ("播放", "views"), ("音乐", "music"), ("视频ID", "video_id"), ("作者抖音号", "author_account"), ("数据来源", "source"), ("备注", "notes")]
RANK_COLS = [("排名", "rank"), ("总分", "score")] + COLLECT_COLS + [("每小时点赞", "likes_per_hour"), ("点赞百分位", "like_pct"), ("评论百分位", "comment_pct"), ("分享百分位", "share_pct"), ("收藏百分位", "favorite_pct"), ("速度百分位", "velocity_pct")]


def _xlsx(path, title, columns, rows):
    wb = Workbook(); ws = wb.active; ws.title = title; ws.append([x[0] for x in columns])
    for row in rows: ws.append([row.get(key, "") for _, key in columns])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]: cell.fill = fill; cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1): ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(path)


def write_collection(output, creators, videos, window, tz, diagnostics):
    output.mkdir(parents=True, exist_ok=True)
    json_path = output / "01_最近7天全部视频.json"; csv_path = output / "01_最近7天全部视频.csv"; xlsx_path = output / "01_最近7天全部视频.xlsx"
    atomic_json(json_path, {"window": window.to_dict(), "creators": [c.to_dict() for c in creators], "videos": [v.to_dict(tz) for v in videos], "diagnostics": diagnostics})
    rows = [_row(v, tz) for v in sorted(videos, key=lambda v: v.create_time or datetime.min.replace(tzinfo=timezone.utc), reverse=True)]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for _, key in COLLECT_COLS]); writer.writeheader(); writer.writerows({key: row.get(key, "") for _, key in COLLECT_COLS} for row in rows)
    _xlsx(xlsx_path, "最近7天全部视频", COLLECT_COLS, rows)
    return {"json": json_path, "csv": csv_path, "xlsx": xlsx_path}


def write_rank(output, ranked, window, tz, config):
    top = ranked[:int(config.get("ranking", {}).get("top_n", 100))]
    json_path = output / "02_最近7天视频Top100.json"; csv_path = output / "02_最近7天视频Top100.csv"; xlsx_path = output / "02_最近7天视频Top100.xlsx"; html_path = output / "02_最近7天视频Top100.html"
    atomic_json(json_path, {"window": window.to_dict(), "videos": [v.to_dict(tz) for v in top]})
    rows = [_row(v, tz) for v in top]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for _, key in RANK_COLS]); writer.writeheader(); writer.writerows({key: row.get(key, "") for _, key in RANK_COLS} for row in rows)
    _xlsx(xlsx_path, "Top100", RANK_COLS, rows)
    cards = "".join(f'<article><b>#{v.rank} {v.final_score:.2f}</b><a href="{html.escape(v.url)}" target="_blank">{html.escape(v.title or "（无标题）")}</a><small>{html.escape(v.creator_name)}｜赞 {v.likes:,}｜分享 {v.shares:,}</small></article>' for v in top)
    html_path.write_text(f'<!doctype html><meta charset="utf-8"><style>body{{font-family:Arial;background:#f3f5f8}}main{{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:12px}}article{{background:white;padding:14px}}a,small{{display:block;margin-top:8px}}</style><h1>指定博主最近7天视频 Top100</h1><p>{html.escape(window.label)}</p><main>{cards}</main>', encoding="utf-8")
    return {"json": json_path, "csv": csv_path, "xlsx": xlsx_path, "html": html_path}


def output_dir(config, window):
    return APP_DIR / str(config.get("output_root", "output/creator_recent7d")) / window.slug


def set_latest(path):
    LATEST_PATH.parent.mkdir(parents=True, exist_ok=True); LATEST_PATH.write_text(str(Path(path).resolve()), encoding="utf-8")


def get_latest(explicit=None):
    path = Path(explicit) if explicit else Path(LATEST_PATH.read_text(encoding="utf-8").strip())
    if not path.exists(): raise FileNotFoundError("没有最近一次采集结果，请先运行功能1")
    return path


def read_collection(path):
    raw = load_json(path)
    return Window.from_dict(raw["window"]), [Creator.from_dict(x) for x in raw.get("creators", [])], [Video.from_dict(x) for x in raw.get("videos", [])], raw.get("diagnostics", {})
