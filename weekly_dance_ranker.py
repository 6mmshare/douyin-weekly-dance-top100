#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""抖音每周单人女性热舞候选 Top100（断点续跑与风控保护版）。

设计原则：
- 使用 Playwright 正常浏览器会话，不绕过验证码或平台签名。
- 监听页面公开返回的 JSON/XHR，并以 DOM 链接作为兜底。
- 对缺失字段如实标记；不伪造发布时间或互动数据。
- 只根据标题、话题、账号自述中的文字信号做内容初筛，
  不从人物外貌推断年龄或性别。
"""

from __future__ import annotations

import argparse
import csv
import dataclasses
import html
import json
import hashlib
import logging
import math
import os
import random
import re
import subprocess
import sys
import time
import traceback
import urllib.parse
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from contextlib import contextmanager
from typing import Any, Callable, Iterable, Iterator, Mapping, Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import BrowserContext, Page, Playwright, Response, TimeoutError as PlaywrightTimeoutError, sync_playwright

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.yaml"
DATA_DIR = APP_DIR / "data"
PROFILE_DIR = DATA_DIR / "profiles"
LOG_DIR = APP_DIR / "logs"
LOG_PATH = LOG_DIR / "weekly_ranker.log"

PLATFORM_LABELS = {"douyin": "抖音", "kuaishou": "快手", "tiktok": "TikTok"}
PLATFORM_HOME = {
    "douyin": "https://www.douyin.com/",
    "kuaishou": "https://www.kuaishou.com/",
    "tiktok": "https://www.tiktok.com/",
}


class CollectorNeedsAttention(RuntimeError):
    """采集需要人工处理时主动停止，避免生成看似成功的空榜。"""


class VerificationRequiredError(CollectorNeedsAttention):
    """后台模式检测到验证码或安全验证。"""


class EmptyDataError(CollectorNeedsAttention):
    """连续搜索没有返回可验证候选，通常是登录失效、风控或页面改版。"""


def enabled_platforms(config: Mapping[str, Any]) -> tuple[str, ...]:
    """返回配置中启用的平台，保持固定展示顺序。"""
    platform_config = config.get("platforms", {})
    return tuple(
        platform
        for platform in ("douyin", "kuaishou", "tiktok")
        if platform_config.get(platform, {}).get("enabled", True)
    )


def report_basename(config: Mapping[str, Any]) -> str:
    total = int(config.get("top_n_per_platform", 50)) * len(enabled_platforms(config))
    return f"weekly_dance_top{total}"

DANCE_TERMS = {
    "舞蹈", "跳舞", "热舞", "独舞", "单人舞", "女团舞", "翻跳",
    "扭腰", "扭胯", "摆胯", "高跟鞋舞", "椅子舞", "慢摇", "对镜热舞", "dance",
    "dancing", "solo dance", "sexy dance", "heels dance", "dance cover", "choreography",
}

SEXY_STYLE_TERMS = {
    "性感", "热辣", "辣妹", "纯欲", "妩媚", "撩人", "御姐", "甜辣", "氛围感", "高级感",
    "扭腰", "扭胯", "摆胯", "腰胯", "身材", "高跟鞋", "椅子舞", "慢摇",
    "sexy", "hot dance", "heels", "body wave", "waacking", "girl crush",
}

FEMALE_TARGET_TERMS = {
    "小姐姐", "美女", "辣妹", "御姐", "女神", "姐姐", "女生", "女性", "girl", "woman",
    "female", "beauty", "queen",
}

SOLO_TARGET_TERMS = {
    "单人", "独舞", "solo", "一个人跳", "个人舞", "对镜", "自拍", "全身舞", "一镜到底",
}

MALE_EXCLUDE_TERMS = {
    "男生跳舞", "男舞者", "男团", "小哥哥跳舞", "正太", "男孩跳舞", "哥哥舞蹈", "猛男舞",
}

GROUP_EXCLUDE_TERMS = {
    "群舞", "多人舞", "双人舞", "三人舞", "团舞", "团播", "群像", "姐妹合跳", "情侣舞",
    "兄妹舞", "合跳", "齐舞", "组合舞", "多人版", "女团",
}

FORMAT_EXCLUDE_TERMS = {
    "舞蹈教程", "教学", "分解教学", "慢动作教学", "镜面教学", "跟练", "动作分解", "教程版",
    "合集", "盘点", "混剪", "影视剪辑", "综艺片段", "reaction", "二创", "搬运", "直播录屏",
}

NON_HUMAN_EXCLUDE_TERMS = {
    "ai舞蹈", "ai美女", "动漫", "动画", "游戏角色", "虚拟人", "数字人", "换脸",
}


GENERIC_TOPIC_STOPWORDS = {
    "舞蹈", "跳舞", "挑战", "热门", "推荐", "教程", "翻跳", "卡点", "热舞", "视频",
    "dance", "dancing", "challenge", "viral", "trend", "trending", "cover", "tutorial",
    "fyp", "foryou", "foryoupage", "tiktok", "douyin", "kuaishou", "girl", "girls",
    "小姐姐", "美女", "女生", "女孩", "女团", "姐姐", "姐妹",
}

COUNT_PATTERN = re.compile(r"(-?\d+(?:\.\d+)?)\s*([万亿千kKmMbB]?)")
HASHTAG_PATTERN = re.compile(r"[#＃]([\w\-\u4e00-\u9fff]+)", re.UNICODE)
CJK_RUN_PATTERN = re.compile(r"[\u4e00-\u9fff]{2,}")
WORD_PATTERN = re.compile(r"[a-zA-Z][a-zA-Z0-9_\-]{1,}")


@dataclass
class VideoRecord:
    platform: str
    video_id: str = ""
    url: str = ""
    title: str = ""
    author_name: str = ""
    author_id: str = ""
    create_time: Optional[datetime] = None
    views: int = 0
    likes: int = 0
    comments: int = 0
    shares: int = 0
    favorites: int = 0
    followers: int = 0
    music: str = ""
    hashtags: list[str] = field(default_factory=list)
    thumbnail: str = ""
    source_keyword: str = ""
    captured_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    data_sources: set[str] = field(default_factory=set)
    data_quality_notes: list[str] = field(default_factory=list)

    dance_relevance: float = 0.0
    target_match_score: float = 0.0
    match_level: str = ""
    female_text_signal: bool = False
    solo_text_signal: bool = False
    sexy_style_signal: bool = False
    exclusion_reason: str = ""
    engagement_rate: float = 0.0
    engagement_basis: str = ""
    velocity_per_hour: float = 0.0
    likes_per_hour: float = 0.0
    cross_platform_score: float = 0.0
    like_percentile: float = 0.0
    comment_percentile: float = 0.0
    share_percentile: float = 0.0
    favorite_percentile: float = 0.0
    share_favorite_percentile: float = 0.0
    engagement_percentile: float = 0.0
    velocity_percentile: float = 0.0
    target_percentile: float = 0.0
    final_score: float = 0.0
    rank: int = 0

    def key(self) -> str:
        if self.video_id:
            return f"{self.platform}:{self.video_id}"
        return f"{self.platform}:{canonicalize_url(self.url)}"

    def local_create_time(self, tz: ZoneInfo) -> Optional[datetime]:
        if not self.create_time:
            return None
        dt = self.create_time
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(tz)

    def to_dict(self, tz: Optional[ZoneInfo] = None) -> dict[str, Any]:
        obj = dataclasses.asdict(self)
        obj["data_sources"] = sorted(self.data_sources)
        if self.create_time:
            obj["create_time"] = (self.local_create_time(tz) if tz else self.create_time).isoformat()
        if self.captured_at:
            obj["captured_at"] = self.captured_at.isoformat()
        return obj


@dataclass
class DateWindow:
    start: datetime
    end: datetime
    mode: str

    @property
    def slug(self) -> str:
        return f"{self.start:%Y-%m-%d}_{(self.end - timedelta(seconds=1)):%Y-%m-%d}"

    @property
    def label(self) -> str:
        return f"{self.start:%Y-%m-%d %H:%M} 至 {self.end:%Y-%m-%d %H:%M}（左闭右开）"


CHECKPOINT_VERSION = 4


def video_record_from_dict(raw: Mapping[str, Any]) -> VideoRecord:
    allowed = {field.name for field in dataclasses.fields(VideoRecord)}
    values = {key: value for key, value in dict(raw).items() if key in allowed}
    values["create_time"] = parse_datetime(values.get("create_time"))
    values["captured_at"] = parse_datetime(values.get("captured_at")) or datetime.now(timezone.utc)
    values["data_sources"] = set(values.get("data_sources") or [])
    values["hashtags"] = list(values.get("hashtags") or [])
    values["data_quality_notes"] = list(values.get("data_quality_notes") or [])
    return VideoRecord(**values)


def atomic_save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    with temp_path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, path)


def collection_fingerprint(platform: str, config: Mapping[str, Any], platform_config: Mapping[str, Any]) -> str:
    browser = config.get("browser", {})
    relevant = {
        "checkpoint_version": CHECKPOINT_VERSION,
        "platform": platform,
        "keywords": platform_config.get("keywords", []),
        "filters": config.get("filters", {}),
        "scoring": config.get("scoring", {}),
        "limits": {
            "max_candidates": browser.get("max_candidates_per_platform"),
            "max_details": browser.get("max_detail_visits_per_platform"),
            "scrolls": browser.get("scrolls_per_keyword"),
        },
    }
    encoded = json.dumps(relevant, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()[:20]


class CollectionCheckpoint:
    def __init__(self, platform: str, window: DateWindow, tz: Any, config: Mapping[str, Any], platform_config: Mapping[str, Any]):
        self.platform = platform
        self.window = window
        self.tz = tz
        self.fingerprint = collection_fingerprint(platform, config, platform_config)
        self.path = DATA_DIR / "checkpoints" / f"{platform}_{window.slug}.json"

    def load(self) -> Optional[dict[str, Any]]:
        if not self.path.exists():
            return None
        try:
            with self.path.open("r", encoding="utf-8") as handle:
                state = json.load(handle)
        except Exception as exc:
            logging.warning("[%s] 断点文件损坏，忽略并重新采集：%s", PLATFORM_LABELS[self.platform], exc)
            return None
        if state.get("version") != CHECKPOINT_VERSION:
            logging.info("[%s] 断点版本已过期，重新采集。", PLATFORM_LABELS[self.platform])
            return None
        if state.get("window_slug") != self.window.slug or state.get("config_fingerprint") != self.fingerprint:
            logging.info("[%s] 搜索词/配置或统计周已变化，旧断点不会复用。", PLATFORM_LABELS[self.platform])
            return None
        return state

    def save(self, records: Iterable[VideoRecord], *, phase: str, next_keyword_index: int, last_keyword_ids: Iterable[str], detail_keys: Iterable[str], completed_detail_keys: Iterable[str], failed_attempts: Mapping[str, int], processed_detail_count: int, note: str = "") -> None:
        records_list = list(records)
        payload = {
            "version": CHECKPOINT_VERSION,
            "platform": self.platform,
            "window_slug": self.window.slug,
            "window_label": self.window.label,
            "config_fingerprint": self.fingerprint,
            "phase": phase,
            "next_keyword_index": int(next_keyword_index),
            "last_keyword_ids": sorted(set(last_keyword_ids)),
            "detail_keys": list(detail_keys),
            "completed_detail_keys": sorted(set(completed_detail_keys)),
            "failed_attempts": {str(k): int(v) for k, v in failed_attempts.items()},
            "processed_detail_count": int(processed_detail_count),
            "record_count": len(records_list),
            "records": [record.to_dict(self.tz) for record in records_list],
            "note": note,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        atomic_save_json(self.path, payload)


class RunLock:
    def __init__(self, name: str, stale_hours: float):
        self.path = DATA_DIR / "locks" / f"{name}.lock"
        self.stale_seconds = max(3600.0, float(stale_hours) * 3600.0)
        self.token = f"{os.getpid()}-{time.time_ns()}"
        self.acquired = False

    def acquire(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            try:
                info = json.loads(self.path.read_text(encoding="utf-8"))
                age = time.time() - self.path.stat().st_mtime
            except Exception:
                info, age = {}, 0
            if age < self.stale_seconds:
                raise RuntimeError(f"另一个任务可能正在运行：{self.path}（PID={info.get('pid', 'unknown')}）")
            logging.warning("发现过期任务锁，自动移除：%s", self.path)
            self.path.unlink(missing_ok=True)
        payload = {"pid": os.getpid(), "token": self.token, "created_at": datetime.now(timezone.utc).isoformat()}
        atomic_save_json(self.path, payload)
        self.acquired = True

    def release(self) -> None:
        if not self.acquired:
            return
        try:
            info = json.loads(self.path.read_text(encoding="utf-8")) if self.path.exists() else {}
            if info.get("token") == self.token:
                self.path.unlink(missing_ok=True)
        except Exception:
            pass
        self.acquired = False

    def __enter__(self) -> "RunLock":
        self.acquire()
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        self.release()


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"配置文件不存在：{path}")
    with path.open("r", encoding="utf-8") as handle:
        config = yaml.safe_load(handle) or {}
    if not isinstance(config, dict):
        raise ValueError("config.yaml 根节点必须是对象。")
    return config


def get_timezone(config: Mapping[str, Any]) -> Any:
    key = str(config.get("timezone", "Asia/Shanghai"))
    try:
        return ZoneInfo(key)
    except (ZoneInfoNotFoundError, ModuleNotFoundError):
        if key in {"Asia/Shanghai", "PRC", "Etc/GMT-8"}:
            logging.warning("系统缺少 tzdata，使用固定 UTC+8；不会影响本工具的中国时区统计。")
            return timezone(timedelta(hours=8), name="Asia/Shanghai")
        logging.warning("找不到时区 %s，使用系统本地时区。", key)
        return datetime.now().astimezone().tzinfo or timezone.utc


def parse_datetime(value: Any) -> Optional[datetime]:
    if value in (None, "", 0, "0"):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        number = float(value)
        if number > 1e12:
            number /= 1000.0
        try:
            return datetime.fromtimestamp(number, tz=timezone.utc)
        except (ValueError, OSError, OverflowError):
            return None
    text = str(value).strip()
    if text.isdigit():
        return parse_datetime(int(text))
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        pass
    patterns = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"]
    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def parse_count(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return max(0, int(value))
    text = str(value).replace(",", "").strip()
    match = COUNT_PATTERN.search(text)
    if not match:
        return 0
    number = float(match.group(1))
    suffix = match.group(2).lower()
    multiplier = {"": 1, "千": 1_000, "k": 1_000, "万": 10_000, "m": 1_000_000, "亿": 100_000_000, "b": 1_000_000_000}.get(suffix, 1)
    return max(0, int(number * multiplier))


def normalize_text(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def canonicalize_url(url: str) -> str:
    if not url:
        return ""
    parts = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
    query = [(k, v) for k, v in query if not k.lower().startswith(("utm_", "share_"))]
    return urllib.parse.urlunsplit((parts.scheme, parts.netloc.lower(), parts.path.rstrip("/"), urllib.parse.urlencode(query), ""))


def normalize_video_url(platform: str, video_id: str, url: str = "") -> str:
    if platform == "douyin" and video_id:
        return f"https://www.douyin.com/video/{video_id}"
    if platform == "kuaishou" and url:
        return canonicalize_url(url)
    if platform == "tiktok" and url:
        return canonicalize_url(url)
    return canonicalize_url(url)


def first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, "", [], {}):
            return value
    return None


def deep_get(obj: Any, *path: Any) -> Any:
    current = obj
    for key in path:
        if isinstance(current, Mapping):
            current = current.get(key)
        elif isinstance(current, list) and isinstance(key, int) and 0 <= key < len(current):
            current = current[key]
        else:
            return None
    return current


def extract_hashtags(title: str, raw_tags: Any = None) -> list[str]:
    tags = {match.group(1).strip() for match in HASHTAG_PATTERN.finditer(title or "") if match.group(1).strip()}
    if isinstance(raw_tags, list):
        for item in raw_tags:
            if isinstance(item, Mapping):
                value = first_value(item.get("hashtag_name"), item.get("tag_name"), item.get("title"), item.get("name"))
            else:
                value = item
            if value:
                tags.add(normalize_text(value).lstrip("#＃"))
    return sorted(tag for tag in tags if tag)


def iter_dicts(value: Any) -> Iterator[dict[str, Any]]:
    if isinstance(value, Mapping):
        yield dict(value)
        for child in value.values():
            yield from iter_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_dicts(child)


def looks_like_douyin_aweme(item: Mapping[str, Any]) -> bool:
    return bool(first_value(item.get("aweme_id"), item.get("item_id"))) and bool(first_value(item.get("desc"), item.get("statistics"), item.get("author"), item.get("video")))


def parse_douyin_item(item: Mapping[str, Any], keyword: str, source: str) -> Optional[VideoRecord]:
    if not looks_like_douyin_aweme(item):
        return None
    video_id = str(first_value(item.get("aweme_id"), item.get("item_id"), item.get("id")) or "")
    statistics = first_value(item.get("statistics"), item.get("stats"), {}) or {}
    author = first_value(item.get("author"), {}) or {}
    video = first_value(item.get("video"), {}) or {}
    title = normalize_text(first_value(item.get("desc"), item.get("title"), item.get("share_info", {}).get("share_title")))
    hashtags = extract_hashtags(title, first_value(item.get("text_extra"), item.get("cha_list")))
    thumbnail = first_value(deep_get(video, "cover", "url_list", 0), deep_get(video, "origin_cover", "url_list", 0), deep_get(video, "dynamic_cover", "url_list", 0), item.get("cover")) or ""
    record = VideoRecord(
        platform="douyin",
        video_id=video_id,
        url=normalize_video_url("douyin", video_id, str(item.get("share_url") or "")),
        title=title,
        author_name=normalize_text(first_value(author.get("nickname"), item.get("author_name"))),
        author_id=str(first_value(author.get("sec_uid"), author.get("uid"), author.get("unique_id"), item.get("author_id")) or ""),
        create_time=parse_datetime(first_value(item.get("create_time"), item.get("createTime"), item.get("timestamp"))),
        views=parse_count(first_value(statistics.get("play_count"), statistics.get("playCount"), statistics.get("view_count"))),
        likes=parse_count(first_value(statistics.get("digg_count"), statistics.get("like_count"), statistics.get("diggCount"))),
        comments=parse_count(first_value(statistics.get("comment_count"), statistics.get("commentCount"))),
        shares=parse_count(first_value(statistics.get("share_count"), statistics.get("shareCount"))),
        favorites=parse_count(first_value(statistics.get("collect_count"), statistics.get("collectCount"))),
        followers=parse_count(first_value(author.get("follower_count"), author.get("mplatform_followers_count"))),
        music=normalize_text(first_value(deep_get(item, "music", "title"), deep_get(item, "music", "music_name"))),
        hashtags=hashtags,
        thumbnail=str(thumbnail),
        source_keyword=keyword,
        data_sources={source},
    )
    if not record.video_id:
        return None
    return record


def looks_like_kuaishou_item(item: Mapping[str, Any]) -> bool:
    return bool(first_value(item.get("photoId"), item.get("photo_id"), item.get("id"))) and bool(first_value(item.get("caption"), item.get("user"), item.get("likeCount"), item.get("viewCount")))


def parse_kuaishou_item(item: Mapping[str, Any], keyword: str, source: str) -> Optional[VideoRecord]:
    if not looks_like_kuaishou_item(item):
        return None
    video_id = str(first_value(item.get("photoId"), item.get("photo_id"), item.get("id")) or "")
    user = first_value(item.get("user"), item.get("author"), {}) or {}
    title = normalize_text(first_value(item.get("caption"), item.get("title"), item.get("description")))
    url = str(first_value(item.get("shareUrl"), item.get("url"), item.get("photoUrl")) or "")
    return VideoRecord(
        platform="kuaishou",
        video_id=video_id,
        url=normalize_video_url("kuaishou", video_id, url),
        title=title,
        author_name=normalize_text(first_value(user.get("name"), user.get("userName"), item.get("userName"))),
        author_id=str(first_value(user.get("id"), user.get("userId"), item.get("userId")) or ""),
        create_time=parse_datetime(first_value(item.get("timestamp"), item.get("createTime"), item.get("create_time"))),
        views=parse_count(first_value(item.get("viewCount"), item.get("playCount"), item.get("browseCount"))),
        likes=parse_count(first_value(item.get("likeCount"), item.get("realLikeCount"))),
        comments=parse_count(first_value(item.get("commentCount"), item.get("comment_count"))),
        shares=parse_count(first_value(item.get("shareCount"), item.get("share_count"))),
        favorites=parse_count(first_value(item.get("collectCount"), item.get("favoriteCount"))),
        followers=parse_count(first_value(user.get("fanCount"), user.get("fansCount"))),
        music=normalize_text(first_value(deep_get(item, "music", "name"), item.get("musicName"))),
        hashtags=extract_hashtags(title, item.get("tagItems")),
        thumbnail=str(first_value(item.get("coverUrl"), item.get("thumbnailUrl"), item.get("cover")) or ""),
        source_keyword=keyword,
        data_sources={source},
    )


def looks_like_tiktok_item(item: Mapping[str, Any]) -> bool:
    return bool(first_value(item.get("id"), item.get("itemId"))) and bool(first_value(item.get("desc"), item.get("stats"), item.get("author"), item.get("video")))


def parse_tiktok_item(item: Mapping[str, Any], keyword: str, source: str) -> Optional[VideoRecord]:
    if not looks_like_tiktok_item(item):
        return None
    video_id = str(first_value(item.get("id"), item.get("itemId")) or "")
    author = first_value(item.get("author"), {}) or {}
    stats = first_value(item.get("stats"), {}) or {}
    author_name = normalize_text(first_value(author.get("uniqueId"), author.get("nickname"), item.get("author")))
    url = str(first_value(item.get("shareUrl"), item.get("webVideoUrl")) or "")
    if not url and author_name and video_id:
        url = f"https://www.tiktok.com/@{author_name}/video/{video_id}"
    title = normalize_text(first_value(item.get("desc"), item.get("title")))
    return VideoRecord(
        platform="tiktok",
        video_id=video_id,
        url=normalize_video_url("tiktok", video_id, url),
        title=title,
        author_name=author_name,
        author_id=str(first_value(author.get("id"), author.get("secUid"), author.get("uniqueId")) or ""),
        create_time=parse_datetime(first_value(item.get("createTime"), item.get("create_time"))),
        views=parse_count(first_value(stats.get("playCount"), stats.get("viewCount"))),
        likes=parse_count(first_value(stats.get("diggCount"), stats.get("likeCount"))),
        comments=parse_count(first_value(stats.get("commentCount"), stats.get("comment_count"))),
        shares=parse_count(first_value(stats.get("shareCount"), stats.get("share_count"))),
        favorites=parse_count(first_value(stats.get("collectCount"), stats.get("saveCount"))),
        followers=parse_count(first_value(author.get("followerCount"), deep_get(item, "authorStats", "followerCount"))),
        music=normalize_text(first_value(deep_get(item, "music", "title"), deep_get(item, "music", "authorName"))),
        hashtags=extract_hashtags(title, item.get("challenges")),
        thumbnail=str(first_value(deep_get(item, "video", "cover"), deep_get(item, "video", "dynamicCover")) or ""),
        source_keyword=keyword,
        data_sources={source},
    )


def parse_json_records(platform: str, payload: Any, keyword: str, source: str) -> list[VideoRecord]:
    parser = {"douyin": parse_douyin_item, "kuaishou": parse_kuaishou_item, "tiktok": parse_tiktok_item}[platform]
    records: list[VideoRecord] = []
    seen: set[str] = set()
    for item in iter_dicts(payload):
        record = parser(item, keyword, source)
        if record and record.key() not in seen:
            seen.add(record.key())
            records.append(record)
    return records


def response_is_search_candidate(platform: str, url: str, current_search: bool = False) -> bool:
    lower = url.lower()
    if platform == "douyin":
        return "search" in lower and any(token in lower for token in ("aweme", "item", "general", "video"))
    if platform == "kuaishou":
        return "search" in lower and any(token in lower for token in ("photo", "feed", "graphql"))
    if platform == "tiktok":
        return "search" in lower and any(token in lower for token in ("item", "video", "general"))
    return current_search and "search" in lower


def response_is_detail_candidate(platform: str, url: str) -> bool:
    lower = url.lower()
    if platform == "douyin":
        return any(token in lower for token in ("aweme/detail", "aweme/v1/web/aweme/detail", "iteminfo"))
    if platform == "kuaishou":
        return any(token in lower for token in ("photo", "detail", "graphql"))
    if platform == "tiktok":
        return any(token in lower for token in ("item/detail", "iteminfo"))
    return False


def merge_record(target: VideoRecord, source: VideoRecord) -> VideoRecord:
    text_fields = ["url", "title", "author_name", "author_id", "music", "thumbnail"]
    numeric_fields = ["views", "likes", "comments", "shares", "favorites", "followers"]
    for field_name in text_fields:
        source_value = getattr(source, field_name)
        if source_value and (not getattr(target, field_name) or len(str(source_value)) > len(str(getattr(target, field_name)))):
            setattr(target, field_name, source_value)
    for field_name in numeric_fields:
        setattr(target, field_name, max(int(getattr(target, field_name) or 0), int(getattr(source, field_name) or 0)))
    if source.create_time and (not target.create_time or source.create_time < target.create_time):
        target.create_time = source.create_time
    target.hashtags = sorted(set(target.hashtags) | set(source.hashtags))
    target.data_sources |= source.data_sources
    target.data_quality_notes = sorted(set(target.data_quality_notes) | set(source.data_quality_notes))
    if source.source_keyword and not target.source_keyword:
        target.source_keyword = source.source_keyword
    return target


def add_records(store: dict[str, VideoRecord], records: Iterable[VideoRecord]) -> None:
    for record in records:
        key = record.key()
        if key in store:
            merge_record(store[key], record)
        else:
            store[key] = record


def record_text(record: VideoRecord) -> str:
    return " ".join([record.title, record.author_name, record.music, " ".join(record.hashtags), record.source_keyword]).lower()


def contains_term(text: str, terms: Iterable[str]) -> bool:
    lower = text.lower()
    return any(str(term).lower() in lower for term in terms if str(term).strip())


def compute_content_signals(record: VideoRecord, config: Mapping[str, Any]) -> None:
    filters = config.get("filters", {})
    text = record_text(record)
    dance_terms = set(DANCE_TERMS)
    sexy_terms = set(SEXY_STYLE_TERMS) | set(filters.get("sexy_style_terms", []) or [])
    female_terms = set(FEMALE_TARGET_TERMS) | set(filters.get("female_text_terms", []) or [])
    solo_terms = set(SOLO_TARGET_TERMS) | set(filters.get("solo_terms", []) or [])
    male_terms = set(MALE_EXCLUDE_TERMS) | set(filters.get("male_terms", []) or [])
    group_terms = set(GROUP_EXCLUDE_TERMS) | set(filters.get("group_terms", []) or [])
    format_terms = set(FORMAT_EXCLUDE_TERMS) | set(filters.get("format_exclude_terms", []) or [])
    non_human_terms = set(NON_HUMAN_EXCLUDE_TERMS) | set(filters.get("non_human_terms", []) or [])
    minor_terms = set(filters.get("minor_terms", []) or [])

    reasons: list[str] = []
    if contains_term(text, male_terms):
        reasons.append("明确男性文本")
    if contains_term(text, group_terms):
        reasons.append("多人/双人/群舞文本")
    if contains_term(text, format_terms):
        reasons.append("教程/合集/搬运等格式")
    if contains_term(text, non_human_terms):
        reasons.append("AI/动漫/虚拟内容")
    if filters.get("exclude_minor_text_signals", True) and contains_term(text, minor_terms):
        reasons.append("明确未成年人文本")

    record.female_text_signal = contains_term(text, female_terms)
    record.solo_text_signal = contains_term(text, solo_terms)
    record.sexy_style_signal = contains_term(text, sexy_terms)

    dance_hits = sum(1 for term in dance_terms if term.lower() in text)
    style_hits = sum(1 for term in sexy_terms if term.lower() in text)
    female_hits = sum(1 for term in female_terms if term.lower() in text)
    solo_hits = sum(1 for term in solo_terms if term.lower() in text)

    record.dance_relevance = min(100.0, dance_hits * 16.0 + style_hits * 8.0 + min(10.0, len(record.hashtags) * 1.5))
    record.target_match_score = min(100.0, dance_hits * 12.0 + style_hits * 11.0 + female_hits * 8.0 + solo_hits * 8.0)
    if record.target_match_score >= 55:
        record.match_level = "A-强匹配"
    elif record.target_match_score >= 28:
        record.match_level = "B-较匹配"
    else:
        record.match_level = "C-宽松候选"
    record.exclusion_reason = "；".join(reasons)


def prefilter_record(record: VideoRecord, config: Mapping[str, Any]) -> bool:
    compute_content_signals(record, config)
    if record.exclusion_reason:
        return False
    filters = config.get("filters", {})
    minimum_target = float(filters.get("minimum_target_match", 8))
    minimum_dance = float(filters.get("minimum_dance_relevance", 12))
    return record.target_match_score >= minimum_target or record.dance_relevance >= minimum_dance


def within_window(record: VideoRecord, window: DateWindow, tz: Any) -> bool:
    local_dt = record.local_create_time(tz)
    return bool(local_dt and window.start <= local_dt < window.end)


def calculate_window(config: Mapping[str, Any], tz: Any, now: Optional[datetime] = None) -> DateWindow:
    now_local = now.astimezone(tz) if now else datetime.now(tz)
    mode = str(config.get("window_mode", "previous_7_complete_days"))
    if mode == "previous_calendar_week":
        this_monday = (now_local - timedelta(days=now_local.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        return DateWindow(this_monday - timedelta(days=7), this_monday, mode)
    today = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    return DateWindow(today - timedelta(days=7), today, mode)


def percentile_map(values: Mapping[str, float]) -> dict[str, float]:
    sorted_items = sorted(values.items(), key=lambda item: item[1])
    n = len(sorted_items)
    if n <= 1:
        return {key: 100.0 for key, _ in sorted_items}
    result: dict[str, float] = {}
    index = 0
    while index < n:
        end = index
        while end + 1 < n and sorted_items[end + 1][1] == sorted_items[index][1]:
            end += 1
        rank = (index + end) / 2.0
        percentile = 100.0 * rank / (n - 1)
        for pos in range(index, end + 1):
            result[sorted_items[pos][0]] = percentile
        index = end + 1
    return result


def score_records(records: list[VideoRecord], config: Mapping[str, Any], tz: Any, captured_at: Optional[datetime] = None) -> list[VideoRecord]:
    captured_at = captured_at or datetime.now(timezone.utc)
    weights = config.get("scoring", {})
    for record in records:
        compute_content_signals(record, config)
        create_time = record.create_time
        if create_time and create_time.tzinfo is None:
            create_time = create_time.replace(tzinfo=timezone.utc)
        hours = max(1.0, (captured_at - create_time.astimezone(timezone.utc)).total_seconds() / 3600.0) if create_time else 0.0
        record.likes_per_hour = record.likes / hours if hours else 0.0
        record.velocity_per_hour = record.likes_per_hour
        if record.followers > 0:
            record.engagement_rate = (record.likes + record.comments + record.shares + record.favorites) / record.followers
            record.engagement_basis = "followers"
        elif record.views > 0:
            record.engagement_rate = (record.likes + record.comments + record.shares + record.favorites) / record.views
            record.engagement_basis = "views"
        else:
            record.engagement_rate = 0.0
            record.engagement_basis = "missing"

    metrics = {
        "like": percentile_map({record.key(): float(record.likes) for record in records}),
        "comment": percentile_map({record.key(): float(record.comments) for record in records}),
        "share": percentile_map({record.key(): float(record.shares) for record in records}),
        "favorite": percentile_map({record.key(): float(record.favorites) for record in records}),
        "velocity": percentile_map({record.key(): float(record.likes_per_hour) for record in records}),
        "target": percentile_map({record.key(): float(record.target_match_score) for record in records}),
    }
    for record in records:
        key = record.key()
        record.like_percentile = metrics["like"].get(key, 0.0)
        record.comment_percentile = metrics["comment"].get(key, 0.0)
        record.share_percentile = metrics["share"].get(key, 0.0)
        record.favorite_percentile = metrics["favorite"].get(key, 0.0)
        record.velocity_percentile = metrics["velocity"].get(key, 0.0)
        record.target_percentile = metrics["target"].get(key, 0.0)
        record.final_score = (
            record.like_percentile * float(weights.get("likes", 0.25))
            + record.share_percentile * float(weights.get("shares", 0.20))
            + record.favorite_percentile * float(weights.get("favorites", 0.15))
            + record.comment_percentile * float(weights.get("comments", 0.10))
            + record.velocity_percentile * float(weights.get("likes_velocity", 0.20))
            + record.target_percentile * float(weights.get("target_match", 0.10))
        )
    ranked = sorted(records, key=lambda record: (record.final_score, record.likes, record.shares, record.favorites), reverse=True)
    for index, record in enumerate(ranked, 1):
        record.rank = index
    return ranked


class BaseCollector:
    def __init__(self, platform: str, config: Mapping[str, Any], platform_config: Mapping[str, Any], tz: Any, window: DateWindow, visible: bool, background: bool = False):
        self.platform = platform
        self.config = config
        self.platform_config = platform_config
        self.tz = tz
        self.window = window
        self.visible = visible
        self.background = background
        self.label = PLATFORM_LABELS[platform]
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.current_keyword = ""
        self.response_records: list[VideoRecord] = []
        self.search_response_count = 0
        self.detail_response_count = 0
        self.debug_dir = APP_DIR / "debug_screenshots"
        self.profile_dir = PROFILE_DIR / platform
        self.profile_dir.mkdir(parents=True, exist_ok=True)
        self.debug_dir.mkdir(parents=True, exist_ok=True)
        self.checkpoint = CollectionCheckpoint(platform, window, tz, config, platform_config)
        self.browser_config = config.get("browser", {})
        self.verification_poll_seconds = max(0.5, float(self.browser_config.get("verification_poll_seconds", 1.5)))
        self.verification_confirm_checks = max(1, int(self.browser_config.get("verification_confirm_checks", 2)))
        self.verification_confirm_interval_seconds = max(0.2, float(self.browser_config.get("verification_confirm_interval_seconds", 0.8)))
        self.verification_min_evidence_score = max(1, int(self.browser_config.get("verification_min_evidence_score", 4)))
        self.verification_wait_timeout_seconds = max(30.0, float(self.browser_config.get("verification_wait_timeout_seconds", 600)))
        self.verification_stable_clear_seconds = max(1.0, float(self.browser_config.get("verification_stable_clear_seconds", 5)))
        self.verification_post_clear_delay_seconds = max(0.0, float(self.browser_config.get("verification_post_clear_delay_seconds", 4)))
        self._verification_active = False
        self._verification_sequence = 0

    def search_url(self, keyword: str) -> str:
        encoded = urllib.parse.quote(keyword)
        if self.platform == "douyin":
            return f"https://www.douyin.com/search/{encoded}?type=video"
        if self.platform == "kuaishou":
            return f"https://www.kuaishou.com/search/video?searchKey={encoded}"
        return f"https://www.tiktok.com/search/video?q={encoded}"

    def launch_context(self, playwright: Playwright) -> BrowserContext:
        args = ["--disable-blink-features=AutomationControlled", "--disable-notifications", "--no-first-run"]
        if self.background:
            x = int(self.browser_config.get("background_window_x", -32000))
            y = int(self.browser_config.get("background_window_y", -32000))
            args.extend([f"--window-position={x},{y}", "--window-size=1440,1000"])
        else:
            args.append("--start-maximized")
        requested_headless = bool(self.browser_config.get("headless", False))
        actual_headless = requested_headless and not self.visible and not self.background
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(self.profile_dir),
            headless=actual_headless,
            locale=str(self.platform_config.get("locale", "zh-CN")),
            viewport={"width": 1440, "height": 1000},
            slow_mo=int(self.browser_config.get("slow_mo_ms", 0)),
            args=args,
        )
        context.set_default_timeout(int(self.browser_config.get("navigation_timeout_ms", 60000)))
        context.on("response", self.on_response)
        return context

    def _safe_visible_count(self, page: Page, selector: str) -> int:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 12)
            visible = 0
            for index in range(count):
                try:
                    element = locator.nth(index)
                    if not element.is_visible(timeout=180):
                        continue
                    box = element.bounding_box(timeout=180)
                    if box and box.get("width", 0) >= 24 and box.get("height", 0) >= 18:
                        visible += 1
                except Exception:
                    continue
            return visible
        except Exception:
            return 0

    def _visible_text_hits(self, page: Page, phrases: Iterable[str]) -> list[str]:
        hits: list[str] = []
        for phrase in phrases:
            try:
                locator = page.get_by_text(phrase, exact=False)
                count = min(locator.count(), 5)
                for index in range(count):
                    element = locator.nth(index)
                    try:
                        if not element.is_visible(timeout=160):
                            continue
                        box = element.bounding_box(timeout=160)
                        if box and box.get("width", 0) >= 30 and box.get("height", 0) >= 16:
                            hits.append(phrase)
                            break
                    except Exception:
                        continue
            except Exception:
                continue
        return hits

    def _normal_content_visible(self, page: Page) -> bool:
        try:
            url = (page.url or "").lower()
        except Exception:
            url = ""
        selectors = [
            'a[href*="/video/"]',
            '[data-e2e*="search-card"]',
            '[data-e2e*="search-video"]',
            '[class*="search-result"] a[href*="/video/"]',
            '[class*="video-card"]',
        ]
        if "/search/" in url or "search?" in url:
            for selector in selectors:
                if self._safe_visible_count(page, selector) >= 2:
                    return True
        if "/video/" in url and self._safe_visible_count(page, "video") >= 1:
            return True
        return False

    def _verification_evidence(self, page: Page) -> tuple[int, list[str]]:
        score = 0
        reasons: list[str] = []
        try:
            url = (page.url or "").lower()
        except Exception:
            url = ""
        if any(token in url for token in ("captcha", "verify", "verification", "security-check")):
            score += 5
            reasons.append("验证URL")

        strong_selectors = [
            'iframe[src*="captcha" i]',
            'iframe[src*="verify" i]',
            'iframe[src*="security" i]',
            '[id*="captcha" i]',
            '[class*="captcha" i]',
            '[id*="verify" i][role="dialog"]',
            '[class*="verify" i][role="dialog"]',
            '[class*="captcha" i][role="dialog"]',
            '[class*="slider" i][class*="verify" i]',
            '[class*="puzzle" i][class*="verify" i]',
            '[aria-modal="true"][class*="verify" i]',
            '[aria-modal="true"][class*="captcha" i]',
        ]
        visible_strong = sum(self._safe_visible_count(page, selector) for selector in strong_selectors)
        if visible_strong:
            score += min(8, visible_strong * 4)
            reasons.append(f"可见验证组件{visible_strong}")

        strong_phrases = [
            "拖动滑块完成拼图",
            "请完成安全验证",
            "请完成下方验证",
            "点击图中",
            "依次点击",
            "请拖动滑块",
            "完成拼图验证",
            "请完成验证后继续",
        ]
        phrase_hits = self._visible_text_hits(page, strong_phrases)
        if phrase_hits:
            score += min(8, len(phrase_hits) * 4)
            reasons.append("可见验证文案:" + "/".join(phrase_hits[:2]))

        overlay_selectors = [
            '[role="dialog"][aria-modal="true"]',
            '[class*="modal-mask" i]',
            '[class*="captcha-mask" i]',
            '[class*="verify-mask" i]',
        ]
        overlay_count = sum(self._safe_visible_count(page, selector) for selector in overlay_selectors)
        if overlay_count and (visible_strong or phrase_hits):
            score += 2
            reasons.append("可见遮罩")

        if self._normal_content_visible(page) and not visible_strong and not phrase_hits:
            return 0, []
        return score, reasons

    def _snapshot_verification(self, page: Page, phase: str) -> Optional[dict[str, Any]]:
        score, reasons = self._verification_evidence(page)
        if score < self.verification_min_evidence_score:
            return None
        return {"page": page, "score": score, "reasons": reasons, "phase": phase, "url": page.url}

    def _detect_verification_once(self, phase: str) -> Optional[dict[str, Any]]:
        if not self.context:
            return None
        for page in list(self.context.pages):
            try:
                if page.is_closed():
                    continue
                hit = self._snapshot_verification(page, phase)
                if hit:
                    return hit
            except Exception:
                continue
        return None

    def detect_verification(self, phase: str = "runtime") -> Optional[dict[str, Any]]:
        first = self._detect_verification_once(phase)
        if not first:
            return None
        for _ in range(self.verification_confirm_checks - 1):
            time.sleep(self.verification_confirm_interval_seconds)
            again = self._detect_verification_once(phase)
            if not again:
                return None
            first = again
        return first

    def _save_verification_screenshot(self, hit: Mapping[str, Any]) -> Optional[Path]:
        if not self.browser_config.get("save_debug_screenshots", True):
            return None
        page = hit.get("page")
        if not isinstance(page, Page):
            return None
        self._verification_sequence += 1
        path = self.debug_dir / f"{self.platform}_verification_{datetime.now():%Y%m%d_%H%M%S}_{self._verification_sequence}.png"
        try:
            page.screenshot(path=str(path), full_page=False)
            return path
        except Exception:
            return None

    def wait_for_verification_if_needed(self, phase: str = "runtime") -> None:
        hit = self.detect_verification(phase)
        if not hit:
            return
        screenshot = self._save_verification_screenshot(hit)
        reason = ", ".join(hit.get("reasons") or ["可见验证组件"])
        if not self.visible:
            extra = f"；截图：{screenshot}" if screenshot else ""
            raise VerificationRequiredError(f"[{self.label}] 检测到可见验证码/安全验证（{reason}，阶段={phase}）{extra}。已保存断点，请运行 run_visible.bat。")

        if not self._verification_active:
            self._verification_active = True
            logging.warning("[%s] 检测到可见验证码/安全验证（%s，阶段=%s）。程序已暂停。", self.label, reason, phase)
            print("\n" + "=" * 68)
            print(f"[{self.label}] 检测到真正可见的验证码或安全验证，程序已暂停。")
            print("请在浏览器中完成验证；无需回到黑框按回车。")
            print("验证消失并稳定后，程序会自动继续当前任务。")
            if screenshot:
                print(f"检测截图：{screenshot}")
            print("=" * 68 + "\n")

        started = time.monotonic()
        clear_since: Optional[float] = None
        last_log = 0.0
        while True:
            current = self._detect_verification_once(phase)
            now_mono = time.monotonic()
            if current:
                clear_since = None
                if now_mono - last_log >= 30:
                    remaining = max(0, int(self.verification_wait_timeout_seconds - (now_mono - started)))
                    logging.info("[%s] 仍在等待人工完成验证，剩余约 %d 秒。", self.label, remaining)
                    last_log = now_mono
            else:
                if clear_since is None:
                    clear_since = now_mono
                    logging.info("[%s] 验证组件已消失，等待稳定 %.0f 秒。", self.label, self.verification_stable_clear_seconds)
                elif now_mono - clear_since >= self.verification_stable_clear_seconds:
                    logging.info("[%s] 验证已完成，%.1f 秒后继续当前任务。", self.label, self.verification_post_clear_delay_seconds)
                    if self.verification_post_clear_delay_seconds:
                        self.monitored_sleep(self.verification_post_clear_delay_seconds, phase=f"{phase}_post_clear", allow_verification=False)
                    self._verification_active = False
                    return
            if now_mono - started >= self.verification_wait_timeout_seconds:
                self._verification_active = False
                raise VerificationRequiredError(f"[{self.label}] 等待人工验证超过 {int(self.verification_wait_timeout_seconds)} 秒。已保存断点，请稍后重新运行 run_visible.bat。")
            time.sleep(self.verification_poll_seconds)

    def monitored_sleep(self, seconds: float, phase: str = "wait", *, allow_verification: bool = True) -> None:
        deadline = time.monotonic() + max(0.0, float(seconds))
        while True:
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                return
            time.sleep(min(self.verification_poll_seconds, remaining))
            if allow_verification:
                self.wait_for_verification_if_needed(phase)

    def wait_between_keywords(self, keyword_index: int, total_keywords: int) -> None:
        if keyword_index >= total_keywords:
            return
        minimum = max(0.0, float(self.browser_config.get("keyword_delay_min_seconds", 8)))
        maximum = max(minimum, float(self.browser_config.get("keyword_delay_max_seconds", 15)))
        delay = random.uniform(minimum, maximum)
        logging.info("[%s] 下一个搜索词前等待 %.1f 秒，并持续检查验证码。", self.label, delay)
        self.monitored_sleep(delay, phase=f"keyword_gap_{keyword_index}")

    def on_response(self, response: Response) -> None:
        try:
            url = response.url
            content_type = (response.headers.get("content-type") or "").lower()
            if "json" not in content_type and not any(token in url.lower() for token in ("api", "graphql", "aweme", "item", "feed", "search")):
                return
            is_search = response_is_search_candidate(self.platform, url, bool(self.current_keyword))
            is_detail = response_is_detail_candidate(self.platform, url)
            if not is_search and not is_detail:
                return
            try:
                payload = response.json()
            except Exception:
                return
            if is_search:
                parsed = parse_json_records(self.platform, payload, self.current_keyword, "search_api")
                if parsed:
                    self.response_records.extend(parsed)
                self.search_response_count += 1
            if is_detail:
                parsed = parse_json_records(self.platform, payload, self.current_keyword, "detail_api")
                if parsed:
                    self.response_records.extend(parsed)
                self.detail_response_count += 1
        except Exception:
            logging.debug("响应解析失败：%s", traceback.format_exc())

    def extract_dom_video_urls(self, page: Page) -> list[str]:
        patterns = {"douyin": "/video/", "kuaishou": "/short-video/", "tiktok": "/video/"}
        token = patterns[self.platform]
        try:
            urls = page.locator(f'a[href*="{token}"]').evaluate_all("elements => elements.map(e => e.href)")
        except Exception:
            return []
        return list(dict.fromkeys(str(url) for url in urls if token in str(url)))

    def records_from_dom_urls(self, urls: Iterable[str], keyword: str) -> list[VideoRecord]:
        records: list[VideoRecord] = []
        for url in urls:
            match = re.search(r"/(?:video|short-video)/(\d+)", url)
            video_id = match.group(1) if match else canonicalize_url(url)
            if not video_id:
                continue
            records.append(VideoRecord(platform=self.platform, video_id=video_id, url=canonicalize_url(url), source_keyword=keyword, data_sources={"search_dom"}))
        return records

    def perform_search(self, keyword: str, retry: bool = False) -> tuple[list[VideoRecord], int]:
        assert self.page is not None
        self.current_keyword = keyword
        self.response_records = []
        before_response_count = self.search_response_count
        self.wait_for_verification_if_needed(f"search_{keyword}_before")
        url = self.search_url(keyword)
        try:
            self.page.goto(url, wait_until="domcontentloaded", timeout=int(self.browser_config.get("navigation_timeout_ms", 60000)))
        except PlaywrightTimeoutError:
            logging.warning("[%s] 搜索页加载超时：%s", self.label, keyword)
        self.wait_for_verification_if_needed(f"search_{keyword}_after_goto")
        self.monitored_sleep(float(self.browser_config.get("request_delay_seconds", 0.7)) + 1.2, phase=f"search_{keyword}_initial_wait")
        self.wait_for_verification_if_needed(f"search_{keyword}_before_scroll")
        scrolls = int(self.browser_config.get("scrolls_per_keyword", 8))
        pause = float(self.browser_config.get("scroll_pause_seconds", 1.5))
        for scroll_index in range(scrolls):
            self.wait_for_verification_if_needed(f"search_{keyword}_scroll_{scroll_index}_before")
            try:
                self.page.mouse.wheel(0, random.randint(850, 1450))
            except Exception:
                pass
            self.monitored_sleep(pause, phase=f"search_{keyword}_scroll_{scroll_index}")
        self.wait_for_verification_if_needed(f"search_{keyword}_after_scroll")
        records = list(self.response_records)
        add_dom = self.records_from_dom_urls(self.extract_dom_video_urls(self.page), keyword)
        records.extend(add_dom)
        unique: dict[str, VideoRecord] = {}
        add_records(unique, records)
        response_delta = self.search_response_count - before_response_count
        return list(unique.values()), response_delta

    def enrich_record(self, record: VideoRecord) -> None:
        assert self.page is not None
        if not record.url:
            return
        self.wait_for_verification_if_needed(f"detail_{record.video_id}_before")
        self.current_keyword = record.source_keyword
        self.response_records = []
        try:
            self.page.goto(record.url, wait_until="domcontentloaded", timeout=int(self.browser_config.get("navigation_timeout_ms", 60000)))
        except PlaywrightTimeoutError:
            record.data_quality_notes.append("详情页加载超时")
        self.wait_for_verification_if_needed(f"detail_{record.video_id}_after_goto")
        self.monitored_sleep(float(self.browser_config.get("detail_wait_seconds", 1.8)), phase=f"detail_{record.video_id}_wait")
        self.wait_for_verification_if_needed(f"detail_{record.video_id}_after_wait")
        for parsed in list(self.response_records):
            if parsed.video_id == record.video_id or len(self.response_records) == 1:
                merge_record(record, parsed)
        if self.platform == "douyin":
            self.enrich_douyin_dom(record)
        self.wait_for_verification_if_needed(f"detail_{record.video_id}_after_dom")

    def enrich_douyin_dom(self, record: VideoRecord) -> None:
        assert self.page is not None
        try:
            body_text = self.page.locator("body").inner_text(timeout=5000)
        except Exception:
            body_text = ""
        if body_text:
            patterns = {
                "likes": [r"(?:点赞|赞)\s*([\d.,万亿kKmM]+)", r"([\d.,万亿kKmM]+)\s*赞"],
                "comments": [r"评论\s*([\d.,万亿kKmM]+)"],
                "favorites": [r"收藏\s*([\d.,万亿kKmM]+)"],
                "shares": [r"分享\s*([\d.,万亿kKmM]+)"],
            }
            for attr, attr_patterns in patterns.items():
                for pattern in attr_patterns:
                    match = re.search(pattern, body_text, flags=re.I)
                    if match:
                        setattr(record, attr, max(getattr(record, attr), parse_count(match.group(1))))
                        record.data_sources.add("detail_dom")
                        break
            date_match = re.search(r"发布时间[:：]?\s*(20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}(?:日)?(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)", body_text)
            if date_match and not record.create_time:
                raw_date = date_match.group(1).replace("年", "-").replace("月", "-").replace("日", "").replace("/", ".")
                record.create_time = parse_datetime(raw_date.replace(".", "-"))
                if record.create_time:
                    record.create_time = record.create_time.replace(tzinfo=self.tz).astimezone(timezone.utc)
        try:
            title = self.page.locator("h1").first.inner_text(timeout=1500)
            if title:
                record.title = max(record.title, normalize_text(title), key=len)
        except Exception:
            pass

    def collect(self) -> tuple[list[VideoRecord], dict[str, Any]]:
        keywords = list(self.platform_config.get("keywords", []))
        if not keywords:
            return [], {"warning": "未配置关键词"}
        max_candidates = int(self.browser_config.get("max_candidates_per_platform", 1200))
        max_details = int(self.browser_config.get("max_detail_visits_per_platform", 250))
        empty_abort_after = int(self.browser_config.get("headless_abort_after_empty_keywords", 4))
        empty_min_candidates = int(self.browser_config.get("headless_min_candidates_before_empty_abort", 8))
        checkpoint_batch_size = max(1, int(self.browser_config.get("checkpoint_batch_size", 20)))
        max_attempts = max(1, int(self.browser_config.get("max_detail_attempts", 2)))
        consecutive_cooldown_after = max(1, int(self.browser_config.get("consecutive_failure_cooldown_after", 4)))
        consecutive_abort_after = max(consecutive_cooldown_after, int(self.browser_config.get("consecutive_failure_abort_after", 8)))
        relevant_store: dict[str, VideoRecord] = {}
        state = self.checkpoint.load()
        next_keyword_index = 0
        last_keyword_ids: set[str] = set()
        detail_keys: list[str] = []
        completed_detail_keys: set[str] = set()
        failed_attempts: dict[str, int] = {}
        processed_detail_count = 0
        phase = "search"
        if state:
            for raw in state.get("records", []):
                try:
                    record = video_record_from_dict(raw)
                    relevant_store[record.key()] = record
                except Exception:
                    continue
            phase = str(state.get("phase") or "search")
            next_keyword_index = int(state.get("next_keyword_index", 0))
            last_keyword_ids = set(state.get("last_keyword_ids", []))
            detail_keys = list(state.get("detail_keys", []))
            completed_detail_keys = set(state.get("completed_detail_keys", []))
            failed_attempts = {str(k): int(v) for k, v in (state.get("failed_attempts") or {}).items()}
            processed_detail_count = int(state.get("processed_detail_count", 0))
            logging.info("[%s] 已恢复断点：阶段=%s，搜索进度=%d/%d，详情已处理=%d，记录=%d", self.label, phase, next_keyword_index, len(keywords), processed_detail_count, len(relevant_store))

        def save_checkpoint(current_phase: str, note: str = "") -> None:
            self.checkpoint.save(
                relevant_store.values(),
                phase=current_phase,
                next_keyword_index=next_keyword_index,
                last_keyword_ids=last_keyword_ids,
                detail_keys=detail_keys,
                completed_detail_keys=completed_detail_keys,
                failed_attempts=failed_attempts,
                processed_detail_count=processed_detail_count,
                note=note,
            )

        empty_streak = 0
        repeated_warning_count = 0
        diagnostics: dict[str, Any] = {"keyword_stats": [], "checkpoint": str(self.checkpoint.path), "resumed": bool(state)}
        with sync_playwright() as playwright:
            self.context = self.launch_context(playwright)
            self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
            try:
                self.wait_for_verification_if_needed("startup")
                if phase == "search":
                    for index in range(next_keyword_index, len(keywords)):
                        keyword = keywords[index]
                        logging.info("[%s] %d/%d 搜索：%s", self.label, index + 1, len(keywords), keyword)
                        self.wait_for_verification_if_needed(f"keyword_{index}_before")
                        records, response_count = self.perform_search(keyword)
                        current_ids = {record.video_id or record.key() for record in records}
                        overlap = len(last_keyword_ids & current_ids) / max(1, len(last_keyword_ids | current_ids)) if last_keyword_ids else 0.0
                        threshold = float(self.browser_config.get("repeated_result_jaccard_threshold", 0.88))
                        if last_keyword_ids and overlap >= threshold:
                            logging.warning("[%s] %s 与上一词结果重复 %.0f%%，改用搜索框重试。", self.label, keyword, overlap * 100)
                            try:
                                search_box = self.page.locator('input[placeholder*="搜索"], input[type="search"]').first
                                search_box.fill(keyword)
                                search_box.press("Enter")
                                self.wait_for_verification_if_needed(f"keyword_{index}_searchbox_after_enter")
                                self.monitored_sleep(3.0, phase=f"keyword_{index}_searchbox_wait")
                                records, response_count = self.perform_search(keyword, retry=True)
                                current_ids = {record.video_id or record.key() for record in records}
                                overlap = len(last_keyword_ids & current_ids) / max(1, len(last_keyword_ids | current_ids)) if last_keyword_ids else 0.0
                            except Exception:
                                logging.debug("搜索框重试失败：%s", traceback.format_exc())
                        before = len(relevant_store)
                        add_records(relevant_store, records)
                        added = len(relevant_store) - before
                        last_keyword_ids = current_ids
                        next_keyword_index = index + 1
                        diagnostics["keyword_stats"].append({"keyword": keyword, "records": len(records), "responses": response_count, "added": added, "overlap_previous": round(overlap, 4)})
                        logging.info("[%s] 本词抓到 %d 条，搜索响应 %d 个，全局去重 %d，本词新增 %d，和上一词重复 %.0f%%", self.label, len(records), response_count, len(relevant_store), added, overlap * 100)
                        if added == 0:
                            empty_streak += 1
                        else:
                            empty_streak = 0
                        if overlap >= threshold:
                            repeated_warning_count += 1
                        save_checkpoint("search", note=f"完成关键词 {keyword}")
                        self.wait_for_verification_if_needed(f"keyword_{index}_after_save")
                        if empty_streak >= empty_abort_after and len(relevant_store) < empty_min_candidates:
                            raise EmptyDataError(f"[{self.label}] 连续 {empty_streak} 个关键词没有产生新的搜索结果，当前全局候选 {len(relevant_store)} 条。已保存搜索断点；请运行 run_visible.bat。")
                        if len(relevant_store) >= max_candidates:
                            logging.info("[%s] 已达到候选上限 %d", self.label, max_candidates)
                            break
                        self.wait_between_keywords(index + 1, len(keywords))
                    phase = "detail"
                    next_keyword_index = len(keywords)

                prefiltered = [record for record in relevant_store.values() if prefilter_record(record, self.config)]
                diagnostics["candidate_count"] = len(relevant_store)
                diagnostics["prefilter_count"] = len(prefiltered)
                diagnostics["excluded_by_text"] = len(relevant_store) - len(prefiltered)
                if not detail_keys:
                    prefiltered.sort(key=lambda record: (record.likes + record.shares * 4 + record.favorites * 3 + record.comments * 2, record.target_match_score), reverse=True)
                    detail_keys = [record.key() for record in prefiltered[:max_details]]
                save_checkpoint("detail", note="详情队列已建立")
                logging.info("[%s] 搜索候选 %d 条，文字初筛可补全 %d 条，排除 %d 条；本轮详情上限 %d", self.label, len(relevant_store), len(prefiltered), len(relevant_store) - len(prefiltered), len(detail_keys))
                pending_keys = [key for key in detail_keys if key not in completed_detail_keys]
                logging.info("[%s] 开始/继续详情补全：总计 %d，已完成 %d，本次待处理 %d；每 %d 条原子保存一次", self.label, len(detail_keys), len(completed_detail_keys), len(pending_keys), checkpoint_batch_size)
                consecutive_failures = 0
                since_last_save = 0
                for key in pending_keys:
                    record = relevant_store.get(key)
                    if not record:
                        completed_detail_keys.add(key)
                        continue
                    success = False
                    last_error = ""
                    for attempt in range(failed_attempts.get(key, 0), max_attempts):
                        try:
                            self.wait_for_verification_if_needed(f"detail_loop_{record.video_id}_before_attempt")
                            self.enrich_record(record)
                            failed_attempts[key] = attempt + 1
                            success = bool(record.title or record.likes or record.comments or record.shares or record.favorites or record.create_time)
                            if success:
                                break
                            last_error = "详情没有返回可验证字段"
                        except VerificationRequiredError:
                            save_checkpoint("detail", note="验证码停止前自动保存")
                            raise
                        except Exception as exc:
                            failed_attempts[key] = attempt + 1
                            last_error = str(exc)
                            backoff = min(float(self.browser_config.get("failure_backoff_max_seconds", 45)), float(self.browser_config.get("failure_backoff_base_seconds", 3)) * (2 ** attempt))
                            logging.warning("[%s] 详情失败 %s，第 %d/%d 次：%s；等待 %.1f 秒", self.label, record.url, attempt + 1, max_attempts, exc, backoff)
                            self.monitored_sleep(backoff + random.uniform(0, 2.0), phase=f"detail_retry_{record.video_id}")
                    if success:
                        consecutive_failures = 0
                    else:
                        consecutive_failures += 1
                        record.data_quality_notes.append(f"详情补全失败：{last_error or '未知错误'}")
                    completed_detail_keys.add(key)
                    processed_detail_count += 1
                    since_last_save += 1
                    if since_last_save >= checkpoint_batch_size:
                        save_checkpoint("detail", note=f"批次保存至 {processed_detail_count}/{len(detail_keys)}")
                        logging.info("[%s] 断点已保存：已处理 %d/%d，成功 %d，最大重试失败 %d", self.label, processed_detail_count, len(detail_keys), processed_detail_count - consecutive_failures, consecutive_failures)
                        since_last_save = 0
                        rest_min = float(self.browser_config.get("batch_rest_min_seconds", 18))
                        rest_max = max(rest_min, float(self.browser_config.get("batch_rest_max_seconds", 40)))
                        rest = random.uniform(rest_min, rest_max)
                        logging.info("[%s] 批次冷却 %.0f 秒，降低连续访问风险。", self.label, rest)
                        self.monitored_sleep(rest, phase=f"detail_batch_rest_{processed_detail_count}")
                    if consecutive_failures == consecutive_cooldown_after:
                        cool_min = float(self.browser_config.get("cooldown_min_seconds", 45))
                        cool_max = max(cool_min, float(self.browser_config.get("cooldown_max_seconds", 90)))
                        cooldown = random.uniform(cool_min, cool_max)
                        logging.warning("[%s] 已连续失败 %d 次，冷却 %.0f 秒。", self.label, consecutive_failures, cooldown)
                        self.monitored_sleep(cooldown, phase=f"detail_failure_cooldown_{processed_detail_count}")
                    if consecutive_failures >= consecutive_abort_after:
                        save_checkpoint("detail", note="连续详情失败，安全停止")
                        raise EmptyDataError(f"[{self.label}] 连续 {consecutive_failures} 条详情加载失败，已保存断点并安全停止。请稍后用 run_visible.bat 检查登录或验证。")
                    delay_min = float(self.browser_config.get("detail_delay_min_seconds", 2.2))
                    delay_max = max(delay_min, float(self.browser_config.get("detail_delay_max_seconds", 4.8)))
                    self.monitored_sleep(random.uniform(delay_min, delay_max), phase=f"detail_gap_{record.video_id}")
                if since_last_save:
                    save_checkpoint("detail", note="详情末尾保存")
                phase = "complete"
                save_checkpoint("complete", note="采集完成")
            except KeyboardInterrupt:
                save_checkpoint(phase, note="用户 Ctrl+C 中断")
                logging.warning("[%s] 收到 Ctrl+C，进度已保存，可下次继续。", self.label)
                raise
            except Exception:
                save_checkpoint(phase, note="异常退出前自动保存")
                raise
            finally:
                try:
                    self.context.close()
                except Exception:
                    pass

        final_records = [record for record in relevant_store.values() if prefilter_record(record, self.config)]
        diagnostics["detail_queue_count"] = len(detail_keys)
        diagnostics["detail_completed_count"] = len(completed_detail_keys)
        diagnostics["detail_failed_attempts"] = sum(1 for value in failed_attempts.values() if value >= max_attempts)
        diagnostics["repeated_warning_count"] = repeated_warning_count
        return final_records, diagnostics


def ensure_directories(config: Mapping[str, Any]) -> None:
    for path in [DATA_DIR, PROFILE_DIR, LOG_DIR, APP_DIR / str(config.get("output_root", "output")), APP_DIR / "debug_screenshots", DATA_DIR / "checkpoints", DATA_DIR / "locks"]:
        path.mkdir(parents=True, exist_ok=True)


def setup_logging() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    handlers = [logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler(sys.stdout)]
    logging.basicConfig(level=logging.INFO, format="%(asctime)s,%(msecs)03d | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S", handlers=handlers, force=True)


def write_csv(path: Path, records: list[VideoRecord], tz: Any) -> None:
    fields = ["排名", "总分", "平台", "视频链接", "标题", "作者", "发布时间", "播放", "点赞", "评论", "分享", "收藏", "发布后每小时点赞", "目标匹配度", "匹配级别", "女性文本信号", "单人文本信号", "性感风格信号", "来源关键词", "数据质量"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for record in records:
            writer.writerow({
                "排名": record.rank,
                "总分": round(record.final_score, 2),
                "平台": PLATFORM_LABELS[record.platform],
                "视频链接": record.url,
                "标题": record.title,
                "作者": record.author_name,
                "发布时间": record.local_create_time(tz).strftime("%Y-%m-%d %H:%M:%S") if record.local_create_time(tz) else "",
                "播放": record.views,
                "点赞": record.likes,
                "评论": record.comments,
                "分享": record.shares,
                "收藏": record.favorites,
                "发布后每小时点赞": round(record.likes_per_hour, 2),
                "目标匹配度": round(record.target_match_score, 2),
                "匹配级别": record.match_level,
                "女性文本信号": "是（仅文字）" if record.female_text_signal else "待人工确认",
                "单人文本信号": "是" if record.solo_text_signal else "待人工确认",
                "性感风格信号": "是" if record.sexy_style_signal else "待人工确认",
                "来源关键词": record.source_keyword,
                "数据质量": "；".join(record.data_quality_notes),
            })


def write_excel(path: Path, platform_records: Mapping[str, list[VideoRecord]], tz: Any, diagnostics: Mapping[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    headers = ["排名", "总分", "封面", "视频/作者", "发布时间", "播放", "点赞", "评论", "分享", "收藏", "每小时点赞", "目标匹配", "匹配级别", "女性文本", "单人文本", "性感风格", "来源关键词", "数据质量", "视频链接"]
    for platform in enabled_platforms({"platforms": {key: {"enabled": key in platform_records} for key in ("douyin", "kuaishou", "tiktok")}}):
        records = platform_records.get(platform, [])
        ws = wb.create_sheet(f"{PLATFORM_LABELS[platform]}Top{len(records)}")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for record in records:
            dt = record.local_create_time(tz)
            ws.append([
                record.rank, round(record.final_score, 2), record.thumbnail,
                f"{record.title}\n{record.author_name}", dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                record.views, record.likes, record.comments, record.shares, record.favorites,
                round(record.likes_per_hour, 2), round(record.target_match_score, 2), record.match_level,
                "是（仅文字）" if record.female_text_signal else "待人工确认",
                "是" if record.solo_text_signal else "待人工确认",
                "是" if record.sexy_style_signal else "待人工确认",
                record.source_keyword, "；".join(record.data_quality_notes), record.url,
            ])
            row = ws.max_row
            ws.cell(row, 4).hyperlink = record.url
            ws.cell(row, 4).style = "Hyperlink"
            ws.cell(row, 19).hyperlink = record.url
            ws.cell(row, 19).style = "Hyperlink"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        widths = [8, 10, 28, 55, 20, 13, 13, 11, 11, 11, 14, 12, 14, 14, 12, 12, 24, 26, 45]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if len(records) >= 2:
            ws.conditional_formatting.add(f"B2:B{len(records)+1}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
        diag_ws = wb.create_sheet(f"{PLATFORM_LABELS[platform]}诊断")
        diag_ws.append(["字段", "内容"])
        for key, value in diagnostics.get(platform, {}).items():
            diag_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)])
        diag_ws.column_dimensions["A"].width = 30
        diag_ws.column_dimensions["B"].width = 120
    wb.save(path)


def html_escape(value: Any) -> str:
    return html.escape(str(value or ""), quote=True)


def write_html(path: Path, platform_records: Mapping[str, list[VideoRecord]], window: DateWindow, config: Mapping[str, Any], tz: Any) -> None:
    all_records: list[VideoRecord] = []
    for platform in enabled_platforms(config):
        all_records.extend(platform_records.get(platform, []))
    cards: list[str] = []
    for index, record in enumerate(all_records, 1):
        dt = record.local_create_time(tz)
        record_id = f"{record.platform}_{record.video_id or index}"
        cover = f'<img src="{html_escape(record.thumbnail)}" alt="封面" loading="lazy">' if record.thumbnail else '<div class="no-cover">暂无封面</div>'
        cards.append(f"""
        <article class="card" data-record-id="{html_escape(record_id)}">
          <label class="keep"><input type="checkbox" checked> 保留候选</label>
          <a class="cover" href="{html_escape(record.url)}" target="_blank" rel="noopener">{cover}</a>
          <div class="content">
            <div class="rank">#{record.rank} <strong>{record.final_score:.2f}</strong> <span>{html_escape(record.match_level)}</span></div>
            <a class="title" href="{html_escape(record.url)}" target="_blank" rel="noopener">{html_escape(record.title or '无标题')}</a>
            <p class="author">{html_escape(record.author_name)}</p>
            <p class="stats">赞 {record.likes:,}　评 {record.comments:,}　分享 {record.shares:,}　收藏 {record.favorites:,}</p>
            <p>每小时点赞 {record.likes_per_hour:,.0f}　匹配 {record.target_match_score:.0f}</p>
            <p>女性词：{'是（仅文字）' if record.female_text_signal else '待人工确认'} / 单人词：{'是' if record.solo_text_signal else '待人工确认'} / 风格：{'是' if record.sexy_style_signal else '待人工确认'}</p>
            <p>{html_escape(dt.strftime('%Y-%m-%d %H:%M:%S') if dt else '发布时间缺失')}｜来源：{html_escape(record.source_keyword)}</p>
            <p class="topics">{html_escape(record.music)} {' '.join('#'+tag for tag in record.hashtags)}</p>
          </div>
        </article>""")
    title = f"抖音 单人女性热舞候选 Top{int(config.get('top_n_per_platform', 100))}"
    html_text = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html_escape(title)}</title>
<style>
body{{margin:0;background:#f4f6fa;color:#14213d;font-family:Arial,"Microsoft YaHei",sans-serif}}header{{background:#19345d;color:#fff;padding:22px}}h1{{margin:0 0 8px}}.note{{margin:18px;background:#fff7d9;border-left:5px solid #d8a20a;padding:13px}}.toolbar{{position:sticky;top:0;z-index:4;background:#edf1f7;padding:12px 18px;display:flex;gap:10px;align-items:center}}button{{border:0;border-radius:5px;padding:10px 14px;background:#245c8d;color:white;font-weight:700}}.secondary{{background:#73829a}}.grid{{padding:18px;display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px}}.card{{position:relative;background:white;border-radius:10px;overflow:hidden;box-shadow:0 2px 10px #0001}}.card.removed{{opacity:.25}}.keep{{position:absolute;z-index:2;top:8px;left:8px;background:#13345dcc;color:white;padding:7px;border-radius:5px}}.cover{{display:block;height:360px;background:#1d2737}}.cover img{{width:100%;height:100%;object-fit:cover}}.no-cover{{height:100%;display:grid;place-items:center;color:#ccc}}.content{{padding:12px}}.rank{{font-size:18px;margin-bottom:8px}}.rank strong{{color:#087f5b}}.rank span{{font-size:12px;border:1px solid #ccd5df;border-radius:10px;padding:3px 6px}}.title{{font-weight:700;color:#075985;text-decoration:none}}p{{margin:7px 0;line-height:1.45}}.author,.topics{{color:#667085;font-size:13px}}
</style></head><body>
<header><h1>{html_escape(title)}</h1><div>统计窗口：{html_escape(window.label)}</div><div>评分：点赞25% + 分享20% + 收藏15% + 评论10% + 每小时点赞20% + 目标匹配度10%</div></header>
<div class="note"><b>用途：</b>这是最多100条“大候选池”，方便人工删到最终Top50。程序不会仅凭外貌自动断言人物性别或成年，仍需打开视频复核。</div>
<div class="toolbar"><button onclick="showAll()">显示全部</button><button class="secondary" onclick="hideRemoved()">隐藏已删除</button><button onclick="exportCsv()">导出保留链接CSV</button><button class="secondary" onclick="resetAll()">重置选择</button><b id="counter"></b></div>
<main class="grid">{''.join(cards)}</main>
<script>
const key='douyin-dance-top100:'+location.pathname;function cards(){{return [...document.querySelectorAll('.card')]}}function load(){{let saved={{}};try{{saved=JSON.parse(localStorage.getItem(key)||'{{}}')}}catch(e){{}}cards().forEach(c=>{{const box=c.querySelector('input');if(c.dataset.recordId in saved)box.checked=!!saved[c.dataset.recordId];box.addEventListener('change',()=>{{save();paint()}})}});paint()}}function save(){{const d={{}};cards().forEach(c=>d[c.dataset.recordId]=c.querySelector('input').checked);localStorage.setItem(key,JSON.stringify(d))}}function paint(){{let n=0;cards().forEach(c=>{{const k=c.querySelector('input').checked;c.classList.toggle('removed',!k);if(k)n++}});counter.textContent=`保留 ${{n}} / ${{cards().length}} 条`}}function showAll(){{cards().forEach(c=>c.style.display='')}}function hideRemoved(){{cards().forEach(c=>c.style.display=c.querySelector('input').checked?'':'none')}}function resetAll(){{cards().forEach(c=>c.querySelector('input').checked=true);save();showAll();paint()}}function exportCsv(){{const rows=[['rank','title','author','url']];cards().filter(c=>c.querySelector('input').checked).forEach(c=>{{const title=c.querySelector('.title');rows.push([c.querySelector('.rank').innerText.split(' ')[0],title.innerText,c.querySelector('.author').innerText,title.href])}});const csv='\ufeff'+rows.map(r=>r.map(v=>'"'+String(v).replaceAll('"','""')+'"').join(',')).join('\n');const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([csv],{{type:'text/csv'}}));a.download='selected_douyin_top.csv';a.click()}}load();
</script></body></html>"""
    path.write_text(html_text, encoding="utf-8")


def write_json(path: Path, records: list[VideoRecord], tz: Any) -> None:
    atomic_save_json(path, [record.to_dict(tz) for record in records])


def generate_outputs(config: Mapping[str, Any], window: DateWindow, platform_records: Mapping[str, list[VideoRecord]], diagnostics: Mapping[str, Any], tz: Any, open_report: bool = False) -> Path:
    output_root = APP_DIR / str(config.get("output_root", "output")) / window.slug
    output_root.mkdir(parents=True, exist_ok=True)
    base = report_basename(config)
    for platform in enabled_platforms(config):
        write_csv(output_root / f"{platform}_top{len(platform_records.get(platform, []))}.csv", platform_records.get(platform, []), tz)
    all_records = [record for platform in enabled_platforms(config) for record in platform_records.get(platform, [])]
    write_json(output_root / "all_candidates.json", all_records, tz)
    write_json(output_root / "diagnostics.json", [], tz)
    atomic_save_json(output_root / "diagnostics.json", diagnostics)
    write_excel(output_root / f"{base}.xlsx", platform_records, tz, diagnostics)
    html_path = output_root / f"{base}.html"
    write_html(html_path, platform_records, window, config, tz)
    if open_report:
        try:
            os.startfile(str(html_path))
        except Exception:
            pass
    logging.info("报告已生成：%s", output_root)
    return output_root


def run_collection(config: Mapping[str, Any], visible: bool, background: bool = False) -> Path:
    ensure_directories(config)
    tz = get_timezone(config)
    window = calculate_window(config, tz)
    logging.info("统计窗口：%s", window.label)
    platform_records: dict[str, list[VideoRecord]] = {}
    diagnostics: dict[str, Any] = {}
    top_n = int(config.get("top_n_per_platform", 100))
    enabled = enabled_platforms(config)
    logging.info("启动采集，关键词 %d 个，模式：%s", sum(len(config.get("platforms", {}).get(p, {}).get("keywords", [])) for p in enabled), "前台" if visible else ("后台兼容（屏幕外正常浏览器）" if background else "无头"))
    stale_hours = float(config.get("browser", {}).get("lock_stale_hours", 12))
    with RunLock("douyin_profile", stale_hours):
        for platform in enabled:
            platform_config = config.get("platforms", {}).get(platform, {})
            collector = BaseCollector(platform, config, platform_config, tz, window, visible=visible, background=background)
            records, diag = collector.collect()
            records = [record for record in records if within_window(record, window, tz) and not record.exclusion_reason]
            ranked = score_records(records, config, tz)[:top_n]
            platform_records[platform] = ranked
            diagnostics[platform] = diag
            logging.info("%s 入榜：%d 条", PLATFORM_LABELS[platform], len(ranked))
    return generate_outputs(config, window, platform_records, diagnostics, tz, open_report=False)


def command_login(config: Mapping[str, Any]) -> None:
    ensure_directories(config)
    with sync_playwright() as playwright:
        for platform in enabled_platforms(config):
            platform_config = config.get("platforms", {}).get(platform, {})
            collector = BaseCollector(platform, config, platform_config, get_timezone(config), calculate_window(config, get_timezone(config)), visible=True)
            context = collector.launch_context(playwright)
            page = context.pages[0] if context.pages else context.new_page()
            page.goto(PLATFORM_HOME[platform], wait_until="domcontentloaded")
            print(f"\n请在浏览器中登录 {PLATFORM_LABELS[platform]}。登录完成并确认首页可用后，回到这里按回车。")
            input()
            context.close()


def command_demo(config: Mapping[str, Any]) -> Path:
    ensure_directories(config)
    tz = get_timezone(config)
    window = calculate_window(config, tz)
    platform_records: dict[str, list[VideoRecord]] = {}
    now = datetime.now(timezone.utc)
    for platform in enabled_platforms(config):
        records: list[VideoRecord] = []
        for index in range(140):
            records.append(VideoRecord(
                platform=platform,
                video_id=f"demo{index+1}",
                url=f"https://example.com/{platform}/video/{index+1}",
                title=f"单人美女热舞演示 #{index+1} #辣妹舞蹈 #扭胯",
                author_name=f"demo_creator_{index+1}",
                create_time=now - timedelta(days=index % 7, hours=index % 20),
                likes=1000 + (140 - index) * 732,
                comments=50 + (index * 17) % 1800,
                shares=35 + (index * 23) % 900,
                favorites=70 + (index * 31) % 1600,
                followers=50_000 + index * 1000,
                source_keyword="单人美女热舞",
                data_sources={"demo"},
            ))
        platform_records[platform] = score_records(records, config, tz)[: int(config.get("top_n_per_platform", 100))]
    return generate_outputs(config, window, platform_records, {p: {"mode": "demo"} for p in platform_records}, tz, open_report=False)


def show_progress(config: Mapping[str, Any]) -> None:
    ensure_directories(config)
    tz = get_timezone(config)
    window = calculate_window(config, tz)
    print(f"统计窗口：{window.label}")
    for platform in enabled_platforms(config):
        path = DATA_DIR / "checkpoints" / f"{platform}_{window.slug}.json"
        print(f"\n[{PLATFORM_LABELS[platform]}]")
        if not path.exists():
            print("尚无断点。")
            continue
        try:
            state = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"断点不可读取：{exc}")
            continue
        print(f"阶段：{state.get('phase')}")
        print(f"搜索进度：{state.get('next_keyword_index', 0)}/{len(config.get('platforms', {}).get(platform, {}).get('keywords', []))}")
        print(f"候选记录：{state.get('record_count', 0)}")
        print(f"详情进度：{len(state.get('completed_detail_keys', []))}/{len(state.get('detail_keys', []))}")
        print(f"最后保存：{state.get('updated_at', '')}")
        print(f"备注：{state.get('note', '')}")


def reset_progress(config: Mapping[str, Any]) -> None:
    ensure_directories(config)
    tz = get_timezone(config)
    window = calculate_window(config, tz)
    removed = 0
    for platform in enabled_platforms(config):
        path = DATA_DIR / "checkpoints" / f"{platform}_{window.slug}.json"
        if path.exists():
            path.unlink()
            removed += 1
            print(f"已删除：{path}")
    print(f"完成，共删除 {removed} 个当前统计周期断点。登录状态未删除。")


def main() -> int:
    parser = argparse.ArgumentParser(description="抖音单人女性热舞候选 Top100")
    parser.add_argument("command", nargs="?", default="run", choices=["run", "login", "demo", "progress", "reset-progress"])
    parser.add_argument("--visible", action="store_true", help="显示浏览器窗口")
    parser.add_argument("--background", action="store_true", help="后台兼容模式：正常浏览器移到屏幕外")
    parser.add_argument("--config", default=str(CONFIG_PATH))
    args = parser.parse_args()
    setup_logging()
    try:
        config = load_config(Path(args.config))
        if args.command == "login":
            command_login(config)
        elif args.command == "demo":
            output = command_demo(config)
            print(f"演示报告：{output}")
        elif args.command == "progress":
            show_progress(config)
        elif args.command == "reset-progress":
            reset_progress(config)
        else:
            output = run_collection(config, visible=args.visible, background=args.background)
            print(f"\n完成：{output}")
        return 0
    except KeyboardInterrupt:
        logging.warning("用户中断，断点已经尽可能保存。")
        return 130
    except CollectorNeedsAttention as exc:
        logging.error("采集已主动停止：%s", exc)
        print(f"\n需要处理：{exc}")
        return 2
    except Exception as exc:
        logging.error("运行失败：%s", exc)
        logging.error(traceback.format_exc())
        print(f"\n运行失败。详细日志：{LOG_PATH}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
